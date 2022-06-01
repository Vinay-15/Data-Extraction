from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import datetime as dt

class restaurant:                                                                              #Creating a class called restaurant
    def __init__(self, link):                                                                  #initializing the link and header file
        self.link = link
        self.hdr = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:84.0) Gecko/20100101 Firefox/84.0',
                    'Accept-Language': 'en-GB,enl;q=0.5', 'Referer': 'https://google.com',
                    'DNT': '1'}                                                                # header used to replicate as a user request rather than python request while accesing websites
        html = requests.get(self.link, headers=self.hdr)
        self.soup = BeautifulSoup(html.text, 'lxml')



    def rest_name(self):                                                                       #to get the name of the restaurant
        if self.link[12:15] == "swi":                                                          #to check if link is of Swiggy
            name = self.soup.find('span', class_='kpkwa')
            return name.text                                                                   #returns name of
        elif self.link[12:15] == "zom":                                                        #to check if link is of Zomato
            name = self.soup.find('h1', class_='sc-7kepeu-0 sc-kafWEX kTxZkY')
            return name.text



    def location(self):                                                                        # to get the location of the restaurant
        if self.link[12:15] == "swi":                                                          #to check if link is of wiggy
            locate = self.soup.find_all('span', class_='_3duMr')
            loc = []                                                                           #assigning empty list
            for l in locate:                                                                   #iterating through all span elements
                loc.append(l.text)                                                             # getting all elements of the location (Home/India/Bangalore/HSR/)
            return loc[-1]                                                                     # the last element is always the location of the restaurant
        elif self.link[12:15] == "zom":                                                        # if the given link is of zomato following lines are executed
            location = self.soup.find_all('span', class_='sc-ks3f96-1 gETRUR')
            loc = []                                                                           #assigning empty list
            for l in location:                                                                 #iterating through all span elements
                loc.append(l.text)                                                             # getting all elements of the location (Home/India/Bangalore/HSR/)
            return (loc[4][:len(loc[4]) - 1])                                                  #usually the 4th element is the location of the restaurant



    def ratings(self):                                                                         # function to get the ratings of the restaurant
        if self.link[12:15] == "swi":                                                          # if the given link is of Swiggy following lines are executed
            Ratings = self.soup.find('div', class_='_1BpLF')
            rate = []                                                                          #assigning empty list
            for r in Ratings:                                                                  #iterating through all span elements
                rate.append(r.text)                                                            # getting all the text from the ratings block
            return (f'{rate[-1][:3]}⭐')                                                      #the last element is always the location of the restaurant
        elif self.link[12:15] == "zom":                                                        # if the given link is of zomato following lines are executed
            Ratings = self.soup.find_all('div', class_='sc-1q7bklc-5 clCBXa')
            rate = []                                                                          # assigning empty list
            for r in Ratings:                                                                  # iterating through all span elements
                rate.append(r.text)                                                            # getting all the text from the ratings block
            return (f'{rate[-1][:3]}⭐')                                                       # the last element is the location sliced till 3 elements



    def sections(self):                         # to get the categories sections (left panel)
        if self.link[12:15] == "swi":                   # if the given link is of Swiggy following lines are executed
            section = self.soup.find_all('h2', class_='M_o7R')         #all the h2 tags are the categories
            self.sec = []                  #creating an empty list called sec
            for sect in section:       #iterating through all h2 tags
                self.sec.append(sect.text)  # making a list of all the h2 tags
            return self.sec
        elif self.link[12:15] == "zom":       # if the given link is of Zomato following lines are executed
            section = self.soup.find_all('p')  # gettiing all the p tags
            self.sec1 = []       #assigning empty list
            sec2 = []  #assigning empty list
            sec3 = []  #assigning empty list
            self.sec = []  #assigning empty list
            for sect in section:   #iterating through all p tags
                sec2.append(sect.text)  # appending all the p tag elements to the list k
            for i in sec2:
                if i != '':  # checking if any empty string elements are present in the list
                    sec3.append(i)  # removing unwanted data from the list and appending it
            for j in range(len(sec3)):
                if sec3[j][-1] == ')':  # Recommended (8) #checking if the last element is ')' then appending it to a list
                    self.sec1.append(sec3[j])
            for p in self.sec1:
                j = p.split('(')       # splitting the category so that it only has the element [Pastas] instead of [Pastas (8)]
                self.sec.append(j[0][:-1])  #slicing till last element to remove the empty spaces in the end
            return self.sec  # Returns the list of all sections



    def check_categories(self):  # to check if the category is present on the website
        self.sections()    #calling another function called sections to get the sections
        a = input("Enter the categories list with a ',' between consecutive elements:").split(
            ',')  # Enter a list of elements separated with a ','
        print('-' * 59)  # to create a table like view in the output
        print('%-50s %-20s' % ('Categories', 'Status'))  #title of the table
        print('-' * 59)
        for i in a:   #iterating through the input categories
            count = 0    #initializing count as 0
            for j in self.sec:
                if i.lower() == j.lower():  # if i=j then count is incremented with 1 else with 0
                    count += 1   #incrementing count if the element is present on the webpage
                else:
                    count += 0
            if (count >= 1):  # check if count>=1 then print that the element is present
                print('%-50s %-20s' % (i, 'Present'))  # %-50s is to allot the item name in those 50 spaces
            else:  #if count is 0 print that the element is absent
                print('%-50s %-20s' % (i, 'Absent'))
        print('-' * 59)   #ending the table view



    def products(self):
        if self.link[12:15] == "swi":  # if the given link is of Swiggy following lines are executed
            products = self.soup.find_all('div', class_='styles_itemName__hLfgz')
            self.productnames = []   #assigning an empty list
            for product in products:   #iterating through all the div tags
                self.productnames.append(product.text)  # holds all the product names
            return self.productnames
        elif self.link[12:15] == "zom":  # if the given link is of Zomato following lines are executed
            products = self.soup.find_all('h4', class_='sc-1s0saks-15 iSmBPS')
            self.productnames = []     #assigning an empty list
            for product in products:    #iterating through all the h4 tags
                self.productnames.append(product.text)  # holds all the product names
            return self.productnames



    def check_products(self):  #function to check if a certain product is present or not
        self.products()    #calling another function called sections to get the sections
        element=self.productnames  #contains product names
        product = input('Enter the products with a , between them :').split(',')  #splitting the product names
        print('-' * 59)  # to create a table-like output
        print('%-50s %-20s' % ('Products', 'Status'))   #title of the table
        print('-' * 59)
        for prod in product:   #iterating through the input list
            count = 0     #initializing count=0
            for ele in element:   #iterating through the actual webpage products
                if prod.lower() == ele.lower():  # if product = ele then count is incremented with 1 else with 0
                    count += 1  #incrementing count if product found
                else:
                    count += 0
            if (count >= 1):  # check if count>=1 then print that the element is present
                print('%-50s %-20s' % (prod, 'Present'))
            else:
                print('%-50s %-20s' % (prod, 'Absent'))
        print('-' * 39)     #end of table



    def menu(self):  # to return a dictionary containing all the sections and products
        if self.link[12:15] == "swi":  #checking if the link is of swiggy
            section = self.soup.find_all('div', class_='_2dS-v')
            category = []   #empty list to hold all categories
            element = []   #empty list to hold all products
            for sec in section:   #iterating through all div tags
                title = sec.find('h2').text  # to get the section names
                products = sec.find_all('h3', class_="styles_itemNameText__3ZmZZ")  # to get all the product names
                pro = []   #empty list to assign products
                for prod in products:  #iterating through the products
                    pro.append(prod.text)   #appending all the products
                category.append(title)   #appending titles
                element.append(pro)  #appending lists of products
            self.menu1 = dict(zip(category, element))  # Joining categories and products make them a dictionary
            return self.menu1
        elif self.link[12:15] == "zom":   #checking if the link is of Zomato
            self.sections()     #calling another function sections()
            self.products()     #calling another function products()
            product_no = []    #assigning empty lists
            prod = []    #assigning empty lists
            for i in self.sec1:  # contains the category list from the previous function (sections)
                k = i.split('(')   #splitting the element
                product_no.append(int(k[-1][:-1]))   #appending all the category numbers
            prod_name = list(self.productnames)  # contains the product names from the previous function (products)
            for j in product_no:     #iterating through all product numbers
                product_list = []    #empty list
                for k in range(j):
                    product_list.append(prod_name[0])   #appending all the product names
                    prod_name.pop(0)     #first element is the section name (to delete that)
                prod.append(product_list)  #appending list of product list
            self.menu1 = dict(zip(self.sec, prod))  # Joining categories and products make them a dictionary
            return self.menu1



    def availability_score(self):    #to get the availability percentage of a certain section
        self.menu()     #calling another function menu()
        inp = input("Enter the section name = ")    #taking the input of a section
        k = int(input("Enter the total number of products of this Category = "))  #input the number of products usually present
        avail = (len(self.menu1[inp]) / k) * 100  #calculating the percentage
        print(f'Availability Score = {avail}%')   #printing the value



'''lnks = ['https://www.zomato.com/bangalore/wowffles-hsr-bangalore/order',
        'https://www.zomato.com/bangalore/prowl-foods-by-tiger-shroff-hsr-bangalore/order',
        'https://www.zomato.com/bangalore/the-thickshake-factory-1-hsr-bangalore/order',
        'https://www.zomato.com/bangalore/wowffles-1-bellandur-bangalore/order',
        'https://www.zomato.com/bangalore/prowl-foods-by-tiger-shroff-sarjapur-road-bangalore/order',
        'https://www.zomato.com/bangalore/the-thickshake-factory-sarjapur-road-bangalore/order',
        'https://www.zomato.com/bangalore/wowffles-koramangala-5th-block-bangalore/order',
        'https://www.zomato.com/bangalore/prowl-foods-by-tiger-shroff-koramangala-5th-block-bangalore/order',
        'https://www.zomato.com/bangalore/the-thickshake-factory-koramangala-5th-block-bangalore/order',
        'https://www.zomato.com/bangalore/wowffles-marathahalli-bangalore/order',
        'https://www.zomato.com/bangalore/prowl-foods-by-tiger-shroff-marathahalli-bangalore/order',
        'https://www.zomato.com/bangalore/the-thickshake-factory-brookefield-bangalore/order',
        'https://www.zomato.com/bangalore/wowffles-1-indiranagar-bangalore/order',
        'https://www.zomato.com/bangalore/prowl-foods-by-tiger-shroff-indiranagar-bangalore/order',
        'https://www.zomato.com/bangalore/the-thick-shake-factory-indiranagar-bangalore/order',
        'https://www.zomato.com/bangalore/wowffles-kalyan-nagar-bangalore/order',
        'https://www.zomato.com/bangalore/prowl-foods-by-tiger-shroff-kalyan-nagar-bangalore/order',
        'https://www.zomato.com/bangalore/the-thickshake-factory-kalyan-nagar-bangalore/order',
        'https://www.swiggy.com/restaurants/wowffles-layout-hsr-bangalore-405914',
        'https://www.swiggy.com/restaurants/prowl-foods-by-tiger-shroff-layout-hsr-bangalore-405913',
        'https://www.swiggy.com/restaurants/the-thick-shake-factory-layout-hsr-bangalore-405912',
        'https://www.swiggy.com/restaurants/wowffles-sarjapur-bellandur-sarjapur-bangalore-405917',
        'https://www.swiggy.com/restaurants/prowl-foods-by-tiger-shroff-sarjapur-bellandur-sarjapur-bangalore-405916',
        'https://www.swiggy.com/restaurants/the-thickshake-factory-opp-central-bellandur-sarjapur-bangalore-56781',
        'https://www.swiggy.com/restaurants/wowffles-koramangala-bangalore-405905',
        'https://www.swiggy.com/restaurants/prowl-foods-by-tiger-shroff-koramangala-bangalore-405904',
        'https://www.swiggy.com/restaurants/the-thick-shake-factory-1st-block-jakkasandra-extension-koramangala-bangalore-26572',
        'https://www.swiggy.com/restaurants/wowffles-marathalli-marathahalli-bangalore-405911',
        'https://www.swiggy.com/restaurants/prowl-foods-by-tiger-shroff-marathalli-marathahalli-bangalore-405910',
        'https://www.swiggy.com/restaurants/the-thick-shake-factory-marathalli-marathahalli-bangalore-405909',
        'https://www.swiggy.com/restaurants/wowffles-indiranagar-bangalore-236428',
        'https://www.swiggy.com/restaurants/prowl-foods-by-tiger-shroff-indira-nagar-indiranagar-bangalore-405907',
        'https://www.swiggy.com/restaurants/the-thick-shake-factory-indira-nagar-indiranagar-bangalore-405906',
        'https://www.swiggy.com/restaurants/wowffles-kalyan-nagar-kammanahalli-kalyan-nagar-bangalore-405920',
        'https://www.swiggy.com/restaurants/prowl-foods-by-tiger-shroff-kalyan-nagar-kammanahalli-kalyan-nagar-bangalore-405919',
        'https://www.swiggy.com/restaurants/the-thickshake-factory-5th-main-kammanahalli-kalyan-nagar-bangalore-34163']'''

lnks=input("Enter a link/list of links seperated with a ',' = ").split(',')
excel = input("Do you want the outputs to be exported to a excel file?  (Y/N) : ")

if excel.lower() == 'n':    #to get the outputs without printing it to excel
    for i in range(len(lnks)):

        obj = restaurant(lnks[i]) #initializing an object to the class restaurant()

        name = obj.rest_name()  # to get the Restaurant name
        print(name)

        locate = obj.location()  # to find the location
        print(locate)

        rate = obj.ratings()  # to find the ratings
        print(rate)

        wp_sections = obj.sections()                                        # to find the categories from the webpage
        print(wp_sections)                                                  #to print webpage category
        obj.check_categories()                                              #to check if certain category is present

        product = obj.products()                                            # to find the products from the webpage
        print(product)                                                      #to print products of the webpage
        obj.check_products()                                                #to check if certain product is present
        

        menu_dict = obj.menu()  # to get a dictionary with categories and products
        if list(menu_dict.keys())[0] == 'Recommended':
            del menu_dict['Recommended']                                    #to delete any dictionary elements
        print(menu_dict)                                                    #to print the dictionary
        #print(len(obj.y['Thick shakes']))                                   #to find the number of elements in any section
        #print(obj.y.keys())                                                 #to print all the section names
        #print(obj.y.values())                                               #to print all the product names

        obj.availability_score()                                             #to get the availability score of any section




elif excel.lower() == 'y':  #to export thr values to an excel file
    wb = Workbook()          #initializing a workbook
    ws = wb.active
    ws.title = "Store Data"         #title of the workbook
    ws.merge_cells('F1:K2')          #merging cells to write the company name in that
    ws['F1'].value = 'TTSF Cloud One'     #company name assigned
    ws['F1'].font = Font(name='Arial', bold=True, size=22, color='00660066', underline="single")  #styling the cell
    ws['F1'].alignment = Alignment(horizontal="center", vertical="center")
    j = '3'    #j contains the row number

    for i in range(len(lnks)):
        obj1 = restaurant(lnks[i])

        if lnks[i][12:15] == "swi":    #checking if the link is of swiggy or zomato
            link = '(Swiggy)'
        else:
            link = '(Zomato)'
        name = obj1.rest_name()  # to get the Restaurant name
        locate = obj1.location()  # to find the location
        rate = obj1.ratings()  # to find the ratings
        menu_dict = obj1.menu()  # to get a dictionary with categories and products
        if list(menu_dict.keys())[0] == 'Recommended':
            del menu_dict['Recommended']   #deleting the recommended section from the menu


      #updating the cells with the above values and also styling them
        ws['A' + j] = name
        ws['A' + j].font = Font(name='Arial', bold=True, size=16, color='00666699', underline='single')
        ws['A' + j].alignment = Alignment(horizontal="center", vertical="center")
        ws['C' + j] = link
        ws['C' + j].font = Font(name='Arial', bold=True, size=10, color='00339966', italic=True)
        ws['C' + j].alignment = Alignment(horizontal="center", vertical="center")
        j = str(int(j) + 1)

        ws['A' + j] = 'Location :'
        ws['A' + j].font = Font(bold=True, size=14, color='00008080')
        ws['B' + j] = locate
        ws['B' + j].font = Font(size=12, color='00666699')
        j = str(int(j) + 1)

        ws['A' + j] = 'Ratings :'
        ws['A' + j].font = Font(bold=True, size=14, color='00008080')
        ws['B' + j] = rate
        ws['B' + j].font = Font(size=12, color='00666699')
        j = str(int(j) + 2)

        ws['A' + j] = 'Categories with Products'
        ws['A' + j].font = Font(bold=True, size=14, color='00008080', italic=True, underline='single')
        j = str(int(j) + 1)

        for key in menu_dict:
            k = key + ' : '
            products = list(menu_dict[key])
            ws.append([k] + products)
            ws['A' + j].font = Font(bold=True, size=12, color='00993300')
            j = str(int(j) + 1)
        j = str(int(j) + 3)

        #using datetime module to use as the file_name
        date = str(dt.datetime.now()).split()[0]         #to get the date
        time = dt.datetime.now().strftime("(%H_%M_%S)")   #to get the time in a formatted way (09_30_25) #hrs_min_sec
        file_name = (date+time)         #assiging date and time to a variable file_name

    ws.column_dimensions['A'].width = 25     #increasing the column width of A for readability
    ws.sheet_view.showGridLines = False    #removing gridlines
    wb.save(file_name+'.xlsx')    #saving the excel file with the specified name

else:
    print('------------- Give a valid input! ----------------- ')    #if given an input other than Y/y or N/n terminate
